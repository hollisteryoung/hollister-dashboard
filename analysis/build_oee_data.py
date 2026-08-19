#!/usr/bin/env python3
"""Derive the OEE figures behind the slide deck and HTML report.

Reads the monthly OEE live file and writes analysis/oee_data.json, the single
source of truth for both deliverables.

The workbook is not tracked on this branch; fetch it with:

    git show 'origin/fabric-native-refresh:Copy of Monthly OEE Live File 2025_July.xlsx' > oee.xlsx

Usage: python3 build_oee_data.py [path/to/oee.xlsx] [-o oee_data.json]
"""

import argparse
import json
import os

import openpyxl

# Reporting periods. April sits with the baseline: at 56.1% it ran below the
# Jan-Mar average and ranked 4th of the seven months, so it is not a step up
# from Q1 and does not belong with the recent run.
BASELINE = ["Jan", "Feb", "Mar", "Apr"]
RECENT = ["May", "Jun", "Jul"]
MONTHS = BASELINE + RECENT
YEAR = "2026"

# Families dropped from the charts and the company average at the user's
# request; still computed so they can be reported separately.
EXCLUDED = {"NewUro", "Rings"}

# Individual (workcell, month) readings dropped from every average, likewise at
# the user's request. The raw value is kept on the machine record so the deck can
# still say what was removed.
EXCLUDED_READINGS = {("BFX1", "Jul")}

FAMILY_LABEL = {
    "2PCAC": "2-Piece Autocoiners",
    "1PCAC": "1-Piece Autocoiners",
    "BIM": "BIM Machines",
    "BFX": "BFX Cells",
    "HolClosed": "Hollister Closed Pouch",
    "HolDrainable": "Hollister Drainable Pouch",
    "DanClosed": "Dansac Closed Pouch",
    "DanDrainable": "Dansac Drainable Pouch",
    "NewUro": "Urostomy (HU) Cells",
    "Rings": "Ring / Extruder",
}

# "OEE% by Month" names each workcell in shop-floor shorthand; "WC Target OEE"
# keys the same asset by its equipment code. This maps one to the other.
WORKCELL_TO_CODE = {
    "1Pc Autocoiner 1 (789)": "83AC1",
    "1Pc Autocoiner 2 (844)": "83AC2",
    "1pc Autocoiner 9 (34753)": "751AC9",
    "1PCAC10": "1AC10",
    "1PCAC11": "1AC11",
    "1PCAC12": "1AC12",
    "Autocoiner #3": "01COINER",
    "Autocoiner #4": "01COIN4",
    "Autocoiner #5": "01COIN5",
    "Autocoiner #6": "01COIN6",
    "Autocoiner #7": "01COIN7",
    "Autocoiner #8": "01COIN8",
    "2Pc Autocoiner 3 (796)": "75AC3",
    "2Pc Autocoiner 4 (843)": "75AC4",
    "2Pc Autocoiner 5 (858)": "75AC5",
    "2Pc Autocoiner 7 (996)": "75AC7",
    "2PCAC06": "2PCAC06",
    "AC08": "2AC08",
    "AC09": "2AC09",
    "AC10": "2AC10",
    "BIM 1 (815)": "90BIM1",
    "BIM 2 (820)": "90BIM2",
    "BIM03": "07BIM",
    "BIM 4 (828)": "90BIM4",
    "BIM 6 (788)": "90BIM6",
    "BIM 7 (798)": "90BIM7",
    "BIM08": "07BIM8",
    "BIM 9 (977)": "90BIM9",
    "BIM 10 (997)": "90BIM10",
    "BIM 11 (991)": "90BIM11",
    "BIM 12": "07BIM12",
    "BIM13": "BIM13",
    "BIM14": "BIM14",
    "BIM15": "BIM15",
    "BIM16": "BIM16",
    "BIM 17": "90BIM17",
    "BIM05": None,          # in service, no target on file
    "BFX1": "07BFX01",
    "PCH01": "PCH01",
    "PCH04": "PCH04",
    "PCH07": "PCH07",
    "PCH15": "PCH15",
    "Kiefel 2": "05KFL02",
    "Kiefel5 (790)": "96K05",
    "Kiefel 3": "05KFL03",
    "Kiefel6 (791)": "96K06",
    "Kiefel8 (793)": "96K08",
    "Kiefel 9": "05KFL09",
    "Keifel11 (978)": "96K11",
    "Keifel12 (979)": "96K12",
    "Kiefel 14": "05KFL14",
    "PCH17": "PCH17",
    "Kiefel 10": "05KFL10",
    "Kiefel 13": "05KFL13",
    "P16": "05PM16",
    "HF11": "05HF11",
    "HF17": "HF17K",
    "HF12": "HF12",
    "HU1 (849)": "78U01",
    "HU2": "01HU2",
    "HU3": "01HU3",
    "Extruder 4 (994)": "92EXT4",
    "EX05": "EXT05",
    "EXT06": "EXT06",
}
FAMILY_OVERRIDE = {"BIM05": "BIM"}


def mean(values):
    """Mean of the readings that exist; None when a workcell never ran."""
    present = [v for v in values if v is not None]
    return sum(present) / len(present) if present else None


def read_targets(wb):
    targets, families = {}, {}
    for row in wb["WC Target OEE"].iter_rows(values_only=True):
        code, target, family = row[1], row[2], row[3]
        if code and family:
            targets[code] = target
            families[code] = family
    return targets, families


def read_machines(wb, targets, families):
    rows = list(wb["OEE% by Month"].iter_rows(values_only=True))
    years, months, measures = rows[4], rows[5], rows[6]

    # Each month carries OEE %, Technical OEE % and TEEP %; keep the first.
    columns = {
        month: i
        for i, (year, month, measure) in enumerate(zip(years, months, measures))
        if year == YEAR and measure and measure.strip() == "OEE %"
    }
    missing = [m for m in MONTHS if m not in columns]
    if missing:
        raise SystemExit(f"{YEAR} columns not found in 'OEE% by Month': {missing}")

    machines = []
    for row in rows[7:]:
        name = row[0]
        if not name or name == "Grand Total":
            continue
        code = WORKCELL_TO_CODE.get(name)
        family = FAMILY_OVERRIDE.get(name) or (families.get(code) if code else None)
        if family is None:
            raise SystemExit(f"no family for workcell {name!r} — update WORKCELL_TO_CODE")
        # '#DIV/0!' means the workcell did not run that month.
        months = {
            m: (row[c] if isinstance(row[c], (int, float)) else None)
            for m, c in columns.items() if m in MONTHS
        }
        dropped = {m: months[m] for m in months if (name, m) in EXCLUDED_READINGS}
        for m in dropped:
            months[m] = None
        machines.append({
            "name": name,
            "code": code,
            "family": family,
            "familyLabel": FAMILY_LABEL[family],
            "target": targets.get(code),
            "months": months,
            "droppedReadings": dropped,
        })
    return machines


def summarise(machines):
    """Family and company roll-ups, as equal-weighted means of monthly OEE."""
    def period(group, period_months):
        return mean([m["months"][k] for m in group for k in period_months])

    families = {}
    for m in machines:
        families.setdefault(m["family"], []).append(m)

    summary = []
    for family, group in families.items():
        summary.append({
            "family": family,
            "label": FAMILY_LABEL[family],
            "n": len(group),
            "excluded": family in EXCLUDED,
            # Workcells with a 0 target have none on file, not a target of zero.
            "target": mean([m["target"] for m in group if m["target"]]),
            "baseline": period(group, BASELINE),
            "recent": period(group, RECENT),
            "months": {k: mean([m["months"][k] for m in group]) for k in MONTHS},
        })
    summary.sort(key=lambda f: -f["recent"])

    included = [m for m in machines if m["family"] not in EXCLUDED]
    company = {
        "n": len(included),
        "target": mean([m["target"] for m in included if m["target"]]),
        "baseline": period(included, BASELINE),
        "recent": period(included, RECENT),
        "months": {k: mean([m["months"][k] for m in included]) for k in MONTHS},
    }
    return summary, company


def main():
    here = os.path.dirname(os.path.abspath(__file__))
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook", nargs="?", default=os.path.join(here, "oee.xlsx"))
    ap.add_argument("-o", "--out", default=os.path.join(here, "oee_data.json"))
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.workbook, data_only=True, read_only=True)
    targets, families = read_targets(wb)
    machines = read_machines(wb, targets, families)
    summary, company = summarise(machines)

    with open(args.out, "w") as fh:
        json.dump({
            "baselineMonths": BASELINE,
            "recentMonths": RECENT,
            "year": YEAR,
            "excludedFamilies": sorted(FAMILY_LABEL[f] for f in EXCLUDED),
            "excludedReadings": [
                {"workcell": n, "month": m, "value": v}
                for mach in machines for m, v in mach["droppedReadings"].items()
                for n in [mach["name"]]
            ],
            "company": company,
            "families": summary,
            "machines": machines,
        }, fh, indent=1)

    print(f"{len(machines)} workcells, {len(summary)} families -> {args.out}")
    print(f"company (excl. {', '.join(sorted(EXCLUDED))}): "
          f"baseline {company['baseline'] * 100:.1f}%  recent {company['recent'] * 100:.1f}%")


if __name__ == "__main__":
    main()
